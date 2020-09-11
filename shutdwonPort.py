import requests
import pprint
import getpass
import wexpect
import openpyxl
import datetime
import time
import re
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill

fileName = 'routerPortInformation.xlsx'

def getDeviceList():
    routerList = list()
    device = requests.get('https://akips11.hsnet.ufl.edu/api-script?password=1r0nM@1d3n;function=web_export_device_list;', verify=False)

    for elem in device.text.splitlines():
        line = elem.split(',')
        temp = list()
        if line[0][1].startswith('R'):
            temp.append(line[0])
            temp.append(line[1].strip())
            routerList.append(temp)
    
    return routerList

def accessJumpBox(username, password):

    print('\n--- Attempting connection to ' + 'IS6 Server... ')
    ssh_newkey = 'Are you sure you want to continue connecting'
    session = wexpect.spawn('ssh ' + username + '@is6.hsnet.ufl.edu')

    idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

    if idx == 0:
        session.sendline('yes')
        idx = session.expect([ssh_newkey, 'word', wexpect.EOF])

        if idx == 0:
            session.sendline(password)
    elif idx == 1:
        session.sendline(password)

    idx = session.expect(['$', wexpect.EOF])

    if idx == 0:
        print("--- Successful Login to JumpBox")
        return session
    else:
        print("--- Terminated program")
        exit()

def accessSwitches(session, switch, username, password):

    session.sendline('ssh ' + switch)

    print('\n------------------------------------------------------')
    print('--- Attempting connection to: ' + switch)
    print('------------------------------------------------------\n')

    session.expect(['word', wexpect.EOF])
    session.sendline(password)
        
    print('--- Success Login to: ', switch)
 
    idx = session.expect(['>', '#', wexpect.EOF])

    if idx == 0:
        session.sendline('en')
        idx = session.expect(['word:', wexpect.EOF])
        session.sendline(password)
        session.expect(['#', wexpect.EOF])
    
    return session

def createExcelFile():
    
    # Excel File Creation
    nowDate = 'Report Date: ' + str(datetime.datetime.now().strftime('%Y-%m-%d'))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Current Port Status'
    
    # Pretty display for the File
    font = Font(bold=True)
    alignment = Alignment(horizontal='center')
    bgColor = PatternFill(fgColor='BFBFBFBF', patternType='solid')
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    ws['A2'] = nowDate
    
    ws['A4'] = 'Hostname'
    ws['A4'].alignment = alignment
    ws['A4'].font = font
    ws['A4'].fill = bgColor
    ws['A4'].border = border

    ws['B4'] = 'IP Address'
    ws['B4'].alignment = alignment
    ws['B4'].font = font
    ws['B4'].fill = bgColor
    ws['B4'].border = border

    ws['C4'] = 'Interface'
    ws['C4'].alignment = alignment
    ws['C4'].font = font  
    ws['C4'].fill = bgColor
    ws['C4'].border = border

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    
    wb.save(fileName)
    wb.close()

def saveExcelFile(deviceList, portList, cellNumber):

    wb = openpyxl.load_workbook(fileName)
    ws = wb.active
    alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

    ws['A' + str(cellNumber)] = deviceList[0]
    ws['A' + str(cellNumber)].alignment = alignment
    ws['A' + str(cellNumber)].border = border

    ws['B' + str(cellNumber)] = deviceList[1]
    ws['B' + str(cellNumber)].alignment = alignment
    ws['B' + str(cellNumber)].border = border

    for elem in portList:
        ws['C' + str(cellNumber)] = elem
        ws['C' + str(cellNumber)].alignment = alignment
        ws['C' + str(cellNumber)].border = border
        
        cellNumber += 1

    wb.save(fileName)

    print('--- Data successfully saved')
    wb.close()

def getPortList(session, switchName):

    portStautsList = list()
    command = 'sh ip int bri | i down'

    session.sendline('term length 0')
    session.expect(['#', wexpect.EOF])
        
    session.sendline(command)
    session.expect(['#', wexpect.EOF])

    data = session.before.splitlines()

    for elem in data[1:][0].split(' '):
        if elem.startswith('Te') or elem.startswith('Gi') or \
            elem.startswith('Tw') or elem.startswith('Hu') or elem.startswith('Fa'):
            portStautsList.append(elem)

    print("--- Complete Gathering Port List Information")
    return portStautsList

def checkAdminDown(session, portList):

    adminDown = list()

    for port in portList:
        
        command = f'sh run int {port}'

        session.sendline(command)
        session.expect(['#', wexpect.EOF])

        for elem in session.before.splitlines()[1:]:
            if elem.strip().startswith('shutdown'):
                adminDown.append(port)
                break
    
    portList = [elem for elem in portList if elem not in adminDown]
    print(adminDown)
    print(portList)
    print("--- Complete Gathering portList to be admin down")
    return portList

if __name__ == '__main__':

    cellNumber = 5
    excludeSwitchList = ['10.4.240.23', '10.4.254.2', '10.4.240.40', 
    '10.4.240.87', '10.4.254.20', '10.4.240.60', '10.4.240.86']

    print()
    print('+-------------------------------------------------------------+')
    print('|    Router physical port security enhancement tool...        |')
    print('|    Make admin down status                                   |')
    print('|    Version 1.0.0                                            |')
    print('|    Compatible with C35xx, C37xx, C38xx, C65XX               |')
    print('|    Nexus 5K, 7K, 9K                                         |')
    print('|    Scripted by Ethan Park, Sep. 2020                        |')
    print('+-------------------------------------------------------------+')
    print()
    username = input("Enter your admin ID ==> ")
    password = getpass.getpass("Enter your password ==> ")
    print()

    switchList = getDeviceList()
    
    createExcelFile()
    portList = list()
    for elem in switchList:
        if elem[1] not in excludeSwitchList:
            session = accessJumpBox(username, password)
            session = accessSwitches(session, elem[1], username, password)
            portList = getPortList(session, elem[0])
            portList = checkAdminDown(session, portList)
            
            saveExcelFile(elem, portList, cellNumber)
            cellNumber += len(portList)
            session.close()